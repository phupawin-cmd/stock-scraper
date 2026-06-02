"""
Excel export module for financial statement data.
Creates well-formatted .xlsx files with separate sheets for each statement.
"""

import os
import re
from typing import Dict, Optional, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR


# ============================================================
# Style definitions
# ============================================================

# Colors
DARK_BLUE = "1F3864"
MED_BLUE = "2F5496"
LIGHT_BLUE = "D6E4F0"
VERY_LIGHT_BLUE = "F2F7FC"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
LIGHT_YELLOW = "FFF8E1"
RED = "CC0000"
DARK_RED = "990000"
GREEN = "006100"
BORDER_GRAY = "BFBFBF"
HEADER_BORDER_COLOR = "1F3864"

# Fonts
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
DATA_FONT = Font(name="Calibri", size=10, color="333333")
BOLD_DATA_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
KEY_ROW_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
PE_FONT = Font(name="Calibri", size=9, italic=True, color="555555")
NEGATIVE_FONT = Font(name="Calibri", size=10, color=RED)

# Fills
HEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
TITLE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ALT_FILL = PatternFill(start_color=VERY_LIGHT_BLUE, end_color=VERY_LIGHT_BLUE, fill_type="solid")
KEY_ROW_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
PE_FILL = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="hair", color=BORDER_GRAY),
    right=Side(style="hair", color=BORDER_GRAY),
    top=Side(style="hair", color=BORDER_GRAY),
    bottom=Side(style="hair", color=BORDER_GRAY),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=HEADER_BORDER_COLOR),
    right=Side(style="thin", color=HEADER_BORDER_COLOR),
    top=Side(style="thin", color=HEADER_BORDER_COLOR),
    bottom=Side(style="medium", color=HEADER_BORDER_COLOR),
)
BOTTOM_BORDER = Border(
    left=Side(style="hair", color=BORDER_GRAY),
    right=Side(style="hair", color=BORDER_GRAY),
    top=Side(style="hair", color=BORDER_GRAY),
    bottom=Side(style="thin", color=DARK_BLUE),
)
PE_BORDER = Border(
    left=Side(style="hair", color=BORDER_GRAY),
    right=Side(style="hair", color=BORDER_GRAY),
    top=Side(style="hair", color=BORDER_GRAY),
    bottom=Side(style="thin", color="CCC29A"),
)

# Number formats
NUMBER_FORMAT = '#,##0.00'
NUMBER_FORMAT_MILLIONS = '#,##0.0'
PERCENT_FORMAT = '0.00%'
DATE_FORMAT = 'YYYY-MM-DD'
MULTIPLIER_FORMAT = '#,##0.00'

# Alignment
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LABEL_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)
NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")
PE_ALIGN = Alignment(horizontal="center", vertical="center")

# ============================================================
# Important rows to highlight (partial match, case-insensitive)
# ============================================================
KEY_ROWS = [
    "revenue", "gross profit", "operating income", "operating expense",
    "net income", "ebitda", "ebit", "eps", "earnings per share",
    "total assets", "total liabilities", "total equity",
    "total current assets", "total current liabilities",
    "total non-current assets", "total non-current liabilities",
    "stockholders' equity", "shareholders' equity",
    "operating cash flow", "free cash flow", "capital expenditure",
    "cash from operations", "cash from investing", "cash from financing",
    "net change in cash", "market cap", "enterprise value",
    "total debt", "net cash",
]

# ============================================================
# Labels & mappings
# ============================================================

STATEMENT_LABELS = {
    "income_statement": "Income Statement",
    "balance_sheet": "Balance Sheet",
    "cash_flow": "Cash Flow Statement",
    "ratios": "Financial Ratios",
    "kpi_metrics": "KPI Metrics",
}

SHEET_NAMES = {
    "income_statement": "Income",
    "balance_sheet": "Balance Sheet",
    "cash_flow": "Cash Flow",
    "ratios": "Ratios",
    "kpi_metrics": "KPI Metrics",
}

PERIOD_LABELS = {
    "Annual": "A",
    "Quarterly": "Q",
}


# ============================================================
# Helper functions
# ============================================================

def ensure_output_dir():
    """Create the output directory if it doesn't exist."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _is_key_row(row_label: str) -> bool:
    """Check if a row label matches an important financial line item."""
    label_lower = row_label.strip().lower()
    for keyword in KEY_ROWS:
        if keyword in label_lower:
            return True
    # Also match rows starting with "Total"
    if label_lower.startswith("total"):
        return True
    return False


def _is_percent_row(row_label: str) -> bool:
    """
    Check if a row should be formatted as percentage based on its label.

    Uses precise keyword matching. Only matches rows whose values
    were originally expressed as percentages on StockAnalysis.com.
    """
    label_lower = row_label.strip().lower()

    # Precise percent-indicating patterns
    percent_patterns = [
        # Margins (Gross Margin, Operating Margin, Profit Margin, EBITDA Margin, etc.)
        "margin",
        # Growth rates (Revenue Growth, EPS Growth, Net Income Growth, etc.)
        "growth",
        # Return ratios (Return on Equity, ROE, ROA, ROIC, ROCE)
        "return on",
        # Yields (Dividend Yield, Earnings Yield, FCF Yield, Buyback Yield)
        "yield",
        # Tax rates
        "effective tax", "tax rate",
        # Payout
        "payout",
        # Ownership percentages
        "owned by",
        # Short interest
        "short %", "short interest",
        # Share changes (Shares Change YoY, Shares Change QoQ)
        "shares change",
        # Dilution
        "dilution",
        # Shareholder return
        "shareholder return",
        # Interest coverage
        "interest coverage",
        # Explicit percent marker
        "(%)", "(%",
    ]

    for pattern in percent_patterns:
        if pattern in label_lower:
            return True

    # Check for ROE/ROA/ROIC/ROCE as standalone acronyms (surrounded by non-letters)
    for acronym in ["roe", "roa", "roic", "roce"]:
        idx = label_lower.find(acronym)
        if idx >= 0:
            # Check character before (if any) is not a letter
            before_ok = idx == 0 or not label_lower[idx - 1].isalpha()
            # Check character after (if any) is not a letter
            after = idx + len(acronym)
            after_ok = after >= len(label_lower) or not label_lower[after].isalpha()
            if before_ok and after_ok:
                return True

    return False


def _is_period_ending(row_label: str) -> bool:
    """Check if row is the Period Ending row."""
    return row_label.strip().lower().startswith("period ending")


def _detect_number_format(row_label: str) -> str:
    """
    Detect the best number format for a row based on its label.
    Returns: 'percent' or 'number'
    """
    if _is_percent_row(row_label):
        return "percent"
    return "number"


def _auto_column_width(ws, min_width=12, max_width=48):
    """Auto-fit column widths based on content."""
    # First column (labels) - wider
    ws.column_dimensions["A"].width = 40

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        if col_letter == "A":
            continue
        max_len = 0
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        # Date/TTM columns need slightly more space
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ============================================================
# Sheet writing
# ============================================================

def _write_dataframe_to_sheet(
    ws, df: pd.DataFrame, sheet_title: str, start_row: int = 1
) -> int:
    """
    Write a DataFrame to a worksheet with professional formatting.

    Features:
    - Freeze panes (header + first column)
    - Auto-filter on header
    - Smart number formatting (% vs absolute)
    - Key row highlighting
    - Negative numbers in red
    - Period Ending row in distinct style
    - Alternating row colors
    - Proper borders and alignment

    Returns the last row number used.
    """
    col_count = len(df.columns) + 1  # +1 for label column
    data_start = start_row + 2  # Row 1 = title, Row 2 = header

    # ---- Row 1: Title ----
    title_cell = ws.cell(row=start_row, column=1, value=sheet_title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=col_count
    )
    ws.row_dimensions[start_row].height = 28

    # ---- Row 2: Header ----
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Line Item")
    for col_idx, col_name in enumerate(df.columns, start=2):
        ws.cell(row=header_row, column=col_idx, value=str(col_name))

    # Header styling
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
    ws.row_dimensions[header_row].height = 32

    # ---- Data rows ----
    for row_idx, (index_val, row_data) in enumerate(df.iterrows()):
        current_row = data_start + row_idx
        row_label = str(index_val)
        is_key = _is_key_row(row_label)
        is_pe = _is_period_ending(row_label)

        # Determine number format for this row
        num_fmt = _detect_number_format(row_label)

        # Column A: Line item label
        label_cell = ws.cell(row=current_row, column=1, value=row_label)
        if is_pe:
            label_cell.font = PE_FONT
        elif is_key:
            label_cell.font = KEY_ROW_FONT
        else:
            label_cell.font = BOLD_DATA_FONT if is_key else DATA_FONT
        label_cell.alignment = LABEL_ALIGN

        # Data columns
        for col_idx, value in enumerate(row_data, start=2):
            if not pd.notna(value):
                continue

            cell = ws.cell(row=current_row, column=col_idx)

            if is_pe:
                # Period Ending: date format
                cell.value = value
                cell.font = PE_FONT
                cell.alignment = PE_ALIGN
                cell.number_format = DATE_FORMAT
            elif num_fmt == "percent":
                cell.value = float(value) if value is not None else None
                cell.font = DATA_FONT
                cell.alignment = NUMBER_ALIGN
                cell.number_format = PERCENT_FORMAT
                # Negative percentages in red
                if isinstance(value, (int, float)) and value < 0:
                    cell.font = NEGATIVE_FONT
            else:
                cell.value = float(value) if value is not None else None
                cell.font = DATA_FONT
                cell.alignment = NUMBER_ALIGN
                cell.number_format = NUMBER_FORMAT
                # Negative values in red
                if isinstance(value, (int, float)) and value < 0:
                    cell.font = NEGATIVE_FONT

        # ---- Row-level styling ----
        for col in range(1, col_count + 1):
            cell = ws.cell(row=current_row, column=col)

            # Borders
            if is_pe:
                cell.border = PE_BORDER
            elif is_key:
                cell.border = BOTTOM_BORDER
            else:
                cell.border = THIN_BORDER

            # Fill (background)
            if is_pe:
                cell.fill = PE_FILL
            elif is_key:
                cell.fill = KEY_ROW_FILL
            elif (row_idx) % 2 == 1:
                cell.fill = ALT_FILL
            else:
                cell.fill = WHITE_FILL

        # Row height
        if is_key:
            ws.row_dimensions[current_row].height = 20
        elif is_pe:
            ws.row_dimensions[current_row].height = 18
        else:
            ws.row_dimensions[current_row].height = 17

    data_end = data_start + len(df) - 1

    # ---- Freeze panes: header + first column ----
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # ---- Auto-filter ----
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{data_end}"

    # ---- Column widths ----
    _auto_column_width(ws)

    # ---- Print settings ----
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return data_end


def export_single_stock(
    symbol: str,
    periods: Dict[str, Dict[str, Optional[pd.DataFrame]]],
    output_dir: str = OUTPUT_DIR,
) -> Optional[str]:
    """
    Export financial statements for a single stock to an Excel file.
    Each statement type + period gets its own sheet.

    Args:
        symbol: Stock ticker symbol
        periods: Dict of period_name -> {statement_type -> DataFrame}
                 e.g. {"Annual": {...}, "Quarterly": {...}}
        output_dir: Directory to save the file

    Returns:
        Path to the created file, or None if no data
    """
    ensure_output_dir()

    # Collect all valid sheets
    all_sheets = {}  # (period_key, stmt_type) -> (sheet_name, df)
    for period_name, statements in periods.items():
        p_key = PERIOD_LABELS.get(period_name, period_name[:1])
        for stmt_type, df in statements.items():
            if df is not None and not df.empty:
                base_name = SHEET_NAMES.get(stmt_type, stmt_type)
                # Sheet name: "Income (A)" or "Income (Q)"
                sheet_name = f"{base_name} ({p_key})"
                # Ensure sheet name fits 31 char limit
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:28] + "...)"
                all_sheets[(period_name, stmt_type)] = (sheet_name, df)

    if not all_sheets:
        print(f"  ⚠️  No data to export for {symbol}")
        return None

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for (period_name, stmt_type), (sheet_name, df) in all_sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        label = STATEMENT_LABELS.get(stmt_type, stmt_type)
        _write_dataframe_to_sheet(ws, df, f"{symbol} - {label} ({period_name})")

    filepath = os.path.join(output_dir, f"{symbol}_financials.xlsx")
    wb.save(filepath)
    print(f"  💾 Saved: {filepath}")
    return filepath


def export_all_stocks(
    all_data: Dict[str, Dict[str, Dict[str, Optional[pd.DataFrame]]]],
    output_dir: str = OUTPUT_DIR,
) -> list:
    """
    Export financial statements for multiple stocks to separate Excel files.

    Args:
        all_data: Nested dict: symbol -> period_name -> statement_type -> DataFrame
        output_dir: Directory to save files

    Returns:
        List of created file paths
    """
    ensure_output_dir()
    created_files = []

    for symbol, periods in all_data.items():
        filepath = export_single_stock(symbol, periods, output_dir)
        if filepath:
            created_files.append(filepath)

    return created_files


def create_summary_file(
    all_data: Dict[str, Dict[str, Dict[str, Optional[pd.DataFrame]]]],
    output_dir: str = OUTPUT_DIR,
) -> Optional[str]:
    """
    Create a summary Excel file with key metrics from all stocks.

    Args:
        all_data: Nested dict: symbol -> period_name -> statement_type -> DataFrame
        output_dir: Directory to save the file

    Returns:
        Path to the summary file
    """
    if not all_data:
        return None

    ensure_output_dir()

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Write header
    ws.cell(row=1, column=1, value="Stock Financial Data Summary").font = Font(
        name="Calibri", size=14, bold=True, color="1F3864"
    )
    ws.merge_cells("A1:K1")

    col = 1
    ws.cell(row=3, column=col, value="Symbol").font = Font(bold=True)
    col += 1
    for period_name in ["Annual", "Quarterly"]:
        for stmt_label in ["Income", "Balance", "Cash Flow", "Ratios", "KPI"]:
            ws.cell(row=3, column=col, value=f"{stmt_label} ({period_name[0]})").font = Font(bold=True)
            col += 1

    row = 4
    stmt_types = ["income_statement", "balance_sheet", "cash_flow", "ratios", "kpi_metrics"]

    for symbol, periods in all_data.items():
        ws.cell(row=row, column=1, value=symbol).font = Font(bold=True)
        col = 2
        for period_name in ["Annual", "Quarterly"]:
            statements = periods.get(period_name, {})
            for stmt_type in stmt_types:
                df = statements.get(stmt_type)
                if df is not None and not df.empty:
                    ws.cell(row=row, column=col, value=f"{df.shape[0]} rows × {df.shape[1]} periods")
                else:
                    ws.cell(row=row, column=col, value="No data")
                col += 1
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 22

    filepath = os.path.join(output_dir, "_summary.xlsx")
    wb.save(filepath)
    print(f"  📋 Summary saved: {filepath}")
    return filepath
